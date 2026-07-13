import io
import logging
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List
from datetime import datetime

from database import get_db
from models import Estimate, Chapter, Position, CompanySettings

router = APIRouter(prefix="/api", tags=["import-export"])
logger = logging.getLogger("smeta")


def get_company_info(db: Session) -> dict:
    settings = db.query(CompanySettings).first()
    if settings:
        return {
            "name": settings.company_name,
            "nif": settings.nif,
            "address": settings.address,
            "phone": settings.phone,
            "email": settings.email,
        }
    return {"name": "", "nif": "", "address": "", "phone": "", "email": ""}


def register_unicode_font():
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import os

    font_paths = [
        "C:/Windows/Fonts/DejaVuSans.ttf",
        "C:/Windows/Fonts/tahoma.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/TTF/DejaVuSans.ttf",
    ]

    for path in font_paths:
        if os.path.exists(path):
            try:
                pdfmetrics.registerFont(TTFont("UnicodeFont", path))
                bold_path = path.replace(".ttf", "bd.ttf").replace("Sans.ttf", "Sans-Bold.ttf")
                if os.path.exists(bold_path):
                    pdfmetrics.registerFont(TTFont("UnicodeFontBold", bold_path))
                else:
                    pdfmetrics.registerFont(TTFont("UnicodeFontBold", path))
                return "UnicodeFont", "UnicodeFontBold"
            except Exception as e:
                logger.warning(f"Could not register font {path}: {e}")

    return "Helvetica", "Helvetica-Bold"


@router.get("/export/{estimate_id}/pdf")
def export_pdf(estimate_id: int, db: Session = Depends(get_db)):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm, mm

    estimate = db.query(Estimate).filter(Estimate.id == estimate_id).first()
    if not estimate:
        raise HTTPException(status_code=404, detail="Estimate not found")

    company = get_company_info(db)
    font_name, font_bold = register_unicode_font()

    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output, pagesize=A4,
        rightMargin=1.5 * cm, leftMargin=1.5 * cm,
        topMargin=1.5 * cm, bottomMargin=2 * cm,
    )

    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "CustomTitle", parent=styles["Title"],
        fontName=font_bold, fontSize=18, spaceAfter=6,
        textColor=colors.HexColor("#1E293B"),
    )
    subtitle_style = ParagraphStyle(
        "Subtitle", parent=styles["Normal"],
        fontName=font_name, fontSize=10, spaceAfter=2,
        textColor=colors.HexColor("#475569"),
    )
    company_style = ParagraphStyle(
        "Company", parent=styles["Normal"],
        fontName=font_name, fontSize=9, spaceAfter=1,
        textColor=colors.HexColor("#334155"),
    )
    normal_style = ParagraphStyle(
        "CustomNormal", parent=styles["Normal"],
        fontName=font_name, fontSize=9,
    )
    small_style = ParagraphStyle(
        "Small", parent=styles["Normal"],
        fontName=font_name, fontSize=8,
        textColor=colors.HexColor("#64748B"),
    )
    total_style = ParagraphStyle(
        "Total", parent=styles["Normal"],
        fontName=font_bold, fontSize=11,
        textColor=colors.HexColor("#1E293B"),
    )

    elements = []

    # Company header (if exists)
    if company["name"]:
        elements.append(Paragraph(company["name"], title_style))
        company_info = []
        if company["nif"]:
            company_info.append(f"NIF/CIF: {company['nif']}")
        if company["address"]:
            company_info.append(company["address"])
        contact_parts = []
        if company["phone"]:
            contact_parts.append(company["phone"])
        if company["email"]:
            contact_parts.append(company["email"])
        if contact_parts:
            company_info.append(" | ".join(contact_parts))
        for line in company_info:
            elements.append(Paragraph(line, company_style))
        elements.append(Spacer(1, 0.5 * cm))
        elements.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#E2E8F0")))
        elements.append(Spacer(1, 0.5 * cm))

    # Estimate title
    elements.append(Paragraph(f"PRESUPUESTO: {estimate.name}", title_style))
    elements.append(Spacer(1, 0.3 * cm))

    # Estimate info
    info_data = []
    if estimate.project_name:
        info_data.append(("Proyecto:", estimate.project_name))
    if estimate.client_name:
        info_data.append(("Cliente:", estimate.client_name))
    if estimate.location:
        info_data.append(("Ubicacion:", estimate.location))
    info_data.append(("Fecha:", datetime.now().strftime("%d/%m/%Y")))
    info_data.append(("Referencia:", f"PPT-{estimate.id:04d}"))

    for label, value in info_data:
        elements.append(Paragraph(f"<b>{label}</b> {value}", normal_style))

    elements.append(Spacer(1, 0.8 * cm))

    # Table header
    table_data = [["Codigo", "Descripcion", "Unidad", "Cant.", "P.U. (EUR)", "Importe (EUR)"]]
    total_general = 0

    for chapter in estimate.chapters:
        chapter_total = sum(p.total_price for p in chapter.positions)
        if chapter_total == 0 and not chapter.positions:
            continue  # Skip empty chapters

        # Chapter header row
        chapter_label = f"{chapter.code} - {chapter.name}" if chapter.name else f"Capitulo {chapter.code}"
        table_data.append([
            Paragraph(f"<b>{chapter_label}</b>", ParagraphStyle("Ch", fontName=font_bold, fontSize=8, textColor=colors.HexColor("#1E40AF"))),
            "", "", "", "",
            Paragraph(f"<b>{chapter_total:,.2f}</b>", ParagraphStyle("ChTotal", fontName=font_bold, fontSize=8, alignment=2))
        ])

        for position in chapter.positions:
            table_data.append([
                position.code,
                Paragraph(position.description, ParagraphStyle("Desc", fontName=font_name, fontSize=8)),
                position.unit,
                f"{position.quantity:,.2f}",
                f"{position.unit_price:,.2f}",
                f"{position.total_price:,.2f}",
            ])

        total_general += chapter_total

    # IVA
    iva_rate = 0.21
    iva_amount = total_general * iva_rate
    total_with_iva = total_general + iva_amount

    table_data.append(["", "", "", "", "Subtotal:", f"{total_general:,.2f}"])
    table_data.append(["", "", "", "", "IVA (21%):", f"{iva_amount:,.2f}"])
    table_data.append(["", "", "", "", Paragraph("<b>TOTAL:</b>", total_style), Paragraph(f"<b>{total_with_iva:,.2f} EUR</b>", total_style)])

    # Column widths
    col_widths = [2 * cm, 6.5 * cm, 1.5 * cm, 1.5 * cm, 2.5 * cm, 2.5 * cm]

    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        # Header
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E40AF")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), font_bold),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ("TOPPADDING", (0, 0), (-1, 0), 6),

        # Data rows
        ("FONTNAME", (0, 1), (-1, -1), font_name),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ALIGN", (3, 1), (5, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 1), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 3),

        # Grid
        ("GRID", (0, 0), (-1, -4), 0.5, colors.HexColor("#CBD5E1")),
        ("LINEBELOW", (0, -4), (-1, -4), 1, colors.HexColor("#94A3B8")),

        # Alternating row colors
        *[("BACKGROUND", (0, i), (-1, i), colors.HexColor("#F8FAFC")) for i in range(2, len(table_data) - 3, 2)],

        # Subtotal/IVA/TOTAL rows
        ("BACKGROUND", (0, -3), (-1, -1), colors.HexColor("#F1F5F9")),
        ("FONTNAME", (0, -3), (-1, -1), font_bold),
        ("LINEABOVE", (0, -3), (-1, -3), 1, colors.HexColor("#94A3B8")),
        ("LINEBELOW", (0, -1), (-1, -1), 2, colors.HexColor("#1E40AF")),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 1 * cm))

    # Footer
    elements.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#E2E8F0")))
    elements.append(Spacer(1, 0.3 * cm))
    elements.append(Paragraph(
        f"Este presupuesto tiene una validez de 30 dias desde la fecha de emision. "
        f"Los precios incluyen materiales y mano de obra salvo indicacion contraria. "
        f"Referencia: PPT-{estimate.id:04d} | Generado por Smeta",
        small_style
    ))

    doc.build(elements)
    output.seek(0)

    filename = f"presupuesto_{estimate.name.replace(' ', '_')}.pdf"
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/export/{estimate_id}/excel")
def export_excel(estimate_id: int, db: Session = Depends(get_db)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers

    estimate = db.query(Estimate).filter(Estimate.id == estimate_id).first()
    if not estimate:
        raise HTTPException(status_code=404, detail="Estimate not found")

    company = get_company_info(db)

    wb = Workbook()
    ws = wb.active
    ws.title = "Presupuesto"

    header_font = Font(bold=True, size=11, color="FFFFFF")
    title_font = Font(bold=True, size=14, color="1E40AF")
    normal_font = Font(size=10)
    bold_font = Font(bold=True, size=10)
    total_font = Font(bold=True, size=12, color="1E40AF")
    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    alt_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    total_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

    row = 1

    # Company info
    if company["name"]:
        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"] = company["name"]
        ws[f"A{row}"].font = title_font
        row += 1
        if company["nif"]:
            ws[f"A{row}"] = f"NIF/CIF: {company['nif']}"
            ws[f"A{row}"].font = normal_font
            row += 1
        if company["address"]:
            ws[f"A{row}"] = company["address"]
            ws[f"A{row}"].font = normal_font
            row += 1
        contact = " | ".join(filter(None, [company.get("phone"), company.get("email")]))
        if contact:
            ws[f"A{row}"] = contact
            ws[f"A{row}"].font = normal_font
            row += 1
        row += 1

    # Estimate info
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = f"PRESUPUESTO: {estimate.name}"
    ws[f"A{row}"].font = title_font
    row += 1

    for label, value in [
        ("Proyecto:", estimate.project_name),
        ("Cliente:", estimate.client_name),
        ("Ubicacion:", estimate.location),
        ("Fecha:", datetime.now().strftime("%d/%m/%Y")),
        ("Referencia:", f"PPT-{estimate.id:04d}"),
    ]:
        if value:
            ws[f"A{row}"] = label
            ws[f"A{row}"].font = bold_font
            ws[f"B{row}"] = value
            ws[f"B{row}"].font = normal_font
            row += 1

    row += 1

    # Table header
    headers = ["Codigo", "Descripcion", "Unidad", "Cantidad", "P.U. (EUR)", "Importe (EUR)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
    row += 1

    total_general = 0

    for chapter in estimate.chapters:
        chapter_total = sum(p.total_price for p in chapter.positions)
        if chapter_total == 0 and not chapter.positions:
            continue

        # Chapter header
        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"] = f"{chapter.code} - {chapter.name}"
        ws[f"A{row}"].font = Font(bold=True, size=10, color="1E40AF")
        for col in range(1, 7):
            ws.cell(row=row, column=col).fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
            ws.cell(row=row, column=col).border = thin_border
        row += 1

        for i, position in enumerate(chapter.positions):
            fill = alt_fill if i % 2 == 0 else PatternFill()
            data = [position.code, position.description, position.unit, position.quantity, position.unit_price, position.total_price]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = normal_font
                cell.border = thin_border
                cell.fill = fill
                if col >= 4:
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal="right")
            row += 1

        total_general += chapter_total

    row += 1

    # Totals
    iva_rate = 0.21
    iva_amount = total_general * iva_rate
    total_with_iva = total_general + iva_amount

    for label, value in [("Subtotal:", total_general), ("IVA (21%):", iva_amount), ("TOTAL:", total_with_iva)]:
        ws.cell(row=row, column=5, value=label).font = bold_font
        ws.cell(row=row, column=5).alignment = Alignment(horizontal="right")
        cell = ws.cell(row=row, column=6, value=value)
        cell.font = total_font if label == "TOTAL:" else bold_font
        cell.number_format = '#,##0.00'
        cell.alignment = Alignment(horizontal="right")
        if label == "TOTAL:":
            cell.fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
            cell.border = Border(top=Side(style="double", color="1E40AF"), bottom=Side(style="double", color="1E40AF"))
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"presupuesto_{estimate.name.replace(' ', '_')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post("/import/excel")
async def import_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    from openpyxl import load_workbook

    content = await file.read()
    wb = load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Empty file")

    headers = [str(h).lower() if h else "" for h in rows[0]]
    column_mapping = {}
    for i, h in enumerate(headers):
        if "codigo" in h or "code" in h:
            column_mapping["code"] = i
        elif "descripcion" in h or "description" in h:
            column_mapping["description"] = i
        elif "unidad" in h or "unit" in h:
            column_mapping["unit"] = i
        elif "cantidad" in h or "quantity" in h:
            column_mapping["quantity"] = i
        elif "precio" in h or "price" in h:
            column_mapping["unit_price"] = i

    if "code" not in column_mapping or "description" not in column_mapping:
        raise HTTPException(status_code=400, detail="Could not identify required columns (code, description)")

    preview = []
    for row in rows[1:11]:
        def get_val(key, default=""):
            idx = column_mapping.get(key)
            if idx is not None and idx < len(row):
                return row[idx] or default
            return default

        preview.append({
            "code": str(get_val("code")),
            "description": str(get_val("description")),
            "unit": str(get_val("unit", "ud")),
            "quantity": float(get_val("quantity", 0) or 0),
            "unit_price": float(get_val("unit_price", 0) or 0),
        })

    return {
        "filename": file.filename,
        "total_rows": len(rows) - 1,
        "columns_found": {k: headers[v] for k, v in column_mapping.items()},
        "preview": preview,
    }
